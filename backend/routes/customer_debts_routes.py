"""
Customer Debts Routes - Extracted from server.py
Debt tracking, payments (oldest-first), summary, Excel export
"""
from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timezone
import uuid


def create_customer_debts_routes(db, get_current_user, get_tenant_admin, require_tenant, CURRENCY="دج") -> dict:
    router = APIRouter(tags=["customer-debts"])

    class CustomerDebtPayment(BaseModel):
        amount: float
        payment_method: str = "cash"
        notes: str = ""

    class SupplierDebtPayment(BaseModel):
        supplier_id: str
        amount: float
        payment_method: str = "cash"

    @router.get("/customers/{customer_id}/debt")
    async def get_customer_debt(customer_id: str, user: dict = Depends(require_tenant)):
        customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
        sales = await db.sales.find({"customer_id": customer_id, "debt_amount": {"$gt": 0}}, {"_id": 0}).sort("created_at", -1).to_list(100)
        payments = await db.debt_payments.find({"customer_id": customer_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
        total_debt = sum(s.get("debt_amount", 0) for s in sales)
        return {"customer_id": customer_id, "customer_name": customer.get("name", ""), "total_debt": total_debt, "unpaid_sales": sales, "payment_history": payments}

    @router.post("/customers/{customer_id}/debt/pay")
    async def pay_customer_debt(customer_id: str, payment: CustomerDebtPayment, user: dict = Depends(require_tenant)):
        customer = await db.customers.find_one({"id": customer_id})
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
        sales = await db.sales.find({"customer_id": customer_id, "debt_amount": {"$gt": 0}}).sort("created_at", 1).to_list(100)
        if not sales:
            raise HTTPException(status_code=400, detail="Customer has no debt")
        remaining_payment = payment.amount
        sales_updated = []
        for sale in sales:
            if remaining_payment <= 0:
                break
            sale_debt = sale.get("debt_amount", 0)
            if sale_debt <= 0:
                continue
            payment_for_sale = min(remaining_payment, sale_debt)
            new_debt = sale_debt - payment_for_sale
            new_paid = sale.get("paid_amount", 0) + payment_for_sale
            await db.sales.update_one({"id": sale["id"]}, {"$set": {"debt_amount": new_debt, "paid_amount": new_paid}})
            remaining_payment -= payment_for_sale
            sales_updated.append({"sale_id": sale["id"], "payment_applied": payment_for_sale, "remaining_debt": new_debt})
        actual_payment = payment.amount - remaining_payment
        payment_record = {"id": str(uuid.uuid4()), "customer_id": customer_id, "customer_name": customer.get("name", ""), "amount": actual_payment, "payment_method": payment.payment_method, "notes": payment.notes, "sales_updated": sales_updated, "created_at": datetime.now(timezone.utc).isoformat(), "created_by": user.get("name", "")}
        await db.debt_payments.insert_one(payment_record)
        if actual_payment > 0:
            await db.customers.update_one({"id": customer_id}, {"$inc": {"total_debt": -actual_payment, "balance": -actual_payment}})
        return {"success": True, "payment_applied": actual_payment, "remaining_from_payment": remaining_payment, "sales_updated": sales_updated}

    @router.post("/supplier-debts/pay")
    async def pay_supplier_debt(payment: SupplierDebtPayment, user: dict = Depends(require_tenant)):
        unpaid = await db.purchases.find({"supplier_id": payment.supplier_id, "remaining": {"$gt": 0}}).sort("created_at", 1).to_list(100)
        if not unpaid:
            raise HTTPException(status_code=400, detail="No outstanding debt for this supplier")
        remaining_payment = payment.amount
        updated_purchases = []
        for purchase in unpaid:
            if remaining_payment <= 0:
                break
            purchase_remaining = purchase["remaining"]
            payment_for_this = min(remaining_payment, purchase_remaining)
            new_paid = purchase["paid_amount"] + payment_for_this
            new_remaining = purchase["total"] - new_paid
            new_status = "paid" if new_remaining <= 0 else "partial"
            await db.purchases.update_one({"id": purchase["id"]}, {"$set": {"paid_amount": new_paid, "remaining": new_remaining, "status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}})
            updated_purchases.append({"purchase_id": purchase["id"], "paid": payment_for_this})
            remaining_payment -= payment_for_this
        supplier = await db.suppliers.find_one({"id": payment.supplier_id})
        if supplier:
            await db.suppliers.update_one({"id": payment.supplier_id}, {"$inc": {"total_purchases": -payment.amount}})
        now = datetime.now(timezone.utc).isoformat()
        await db.transactions.insert_one({"id": str(uuid.uuid4()), "type": "expense", "box": payment.payment_method, "amount": -payment.amount, "balance_after": 0, "description": f"سداد دين مورد - {supplier['name'] if supplier else payment.supplier_id}", "created_at": now})
        await db.cash_boxes.update_one({"id": payment.payment_method}, {"$inc": {"balance": -payment.amount}})
        return {"message": "Payment recorded successfully", "amount_paid": payment.amount, "updated_purchases": updated_purchases}

    @router.get("/debts/summary")
    async def get_debts_summary(user: dict = Depends(require_tenant)):
        pipeline = [{"$match": {"debt_amount": {"$gt": 0}}}, {"$group": {"_id": "$customer_id", "total_debt": {"$sum": "$debt_amount"}, "sales_count": {"$sum": 1}}}]
        debts_by_customer = await db.sales.aggregate(pipeline).to_list(1000)
        result = []
        for debt in debts_by_customer:
            customer = await db.customers.find_one({"id": debt["_id"]}, {"_id": 0, "name": 1, "phone": 1})
            if customer:
                result.append({"customer_id": debt["_id"], "customer_name": customer.get("name", "Unknown"), "customer_phone": customer.get("phone", ""), "total_debt": debt["total_debt"], "sales_count": debt["sales_count"]})
        total_outstanding = sum(d["total_debt"] for d in result)
        return {"total_outstanding": total_outstanding, "customers_with_debt": len(result), "debts": sorted(result, key=lambda x: x["total_debt"], reverse=True)}

    @router.get("/debts/export")
    async def export_debts_to_excel(user: dict = Depends(require_tenant)):
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from fastapi.responses import StreamingResponse
        pipeline = [{"$match": {"debt_amount": {"$gt": 0}}}, {"$group": {"_id": "$customer_id", "total_debt": {"$sum": "$debt_amount"}, "sales_count": {"$sum": 1}}}]
        debts_by_customer = await db.sales.aggregate(pipeline).to_list(1000)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Debts"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        headers = ["#", "اسم الزبون", "رقم الهاتف", "عدد الفواتير", f"إجمالي الدين ({CURRENCY})"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        row_num = 2
        total_debt = 0
        for idx, debt in enumerate(debts_by_customer, 1):
            customer = await db.customers.find_one({"id": debt["_id"]}, {"_id": 0})
            if not customer:
                continue
            ws.cell(row=row_num, column=1, value=idx).border = border
            ws.cell(row=row_num, column=2, value=customer.get("name", "")).border = border
            ws.cell(row=row_num, column=3, value=customer.get("phone", "")).border = border
            ws.cell(row=row_num, column=4, value=debt["sales_count"]).border = border
            cell = ws.cell(row=row_num, column=5, value=debt["total_debt"])
            cell.border = border
            cell.number_format = '#,##0.00'
            total_debt += debt["total_debt"]
            row_num += 1
        ws.cell(row=row_num, column=4, value="الإجمالي:").font = Font(bold=True)
        total_cell = ws.cell(row=row_num, column=5, value=total_debt)
        total_cell.font = Font(bold=True, color="FF0000")
        total_cell.number_format = '#,##0.00'
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.sheet_view.rightToLeft = True
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=debts_{datetime.now().strftime('%Y%m%d')}.xlsx"})

    return router
