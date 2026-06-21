"""
Data Import/Export Routes - CSV and Excel support
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from typing import Optional
import csv
import io
import json
from datetime import datetime, timezone
from bson import ObjectId
from openpyxl import Workbook, load_workbook
from config.database import get_tenant_db

def create_import_export_routes(db, get_current_user) -> dict:
    router = APIRouter(prefix="/data", tags=["data-import-export"])

    EXPORTABLE_COLLECTIONS = {
        "products": {
            "fields": ["name_ar", "name_en", "barcode", "article_code", "retail_price", "wholesale_price", "purchase_price", "quantity", "min_stock", "category", "family_id", "unit", "tax_rate"],
            "label_ar": "المنتجات",
            "label_fr": "Products"
        },
        "customers": {
            "fields": ["name", "phone", "email", "address", "city", "wilaya", "notes", "family_id"],
            "label_ar": "الزبائن",
            "label_fr": "Customers"
        },
        "suppliers": {
            "fields": ["name", "phone", "email", "address", "city", "company", "tax_id", "notes"],
            "label_ar": "الموردين",
            "label_fr": "Suppliers"
        },
        "employees": {
            "fields": ["name", "phone", "email", "position", "salary", "hire_date", "notes"],
            "label_ar": "الموظفين",
            "label_fr": "Employees"
        },
        "sales": {
            "fields": ["invoice_number", "customer_name", "total", "discount", "payment_method", "payment_type", "status", "created_at", "note"],
            "label_ar": "المبيعات",
            "label_fr": "Sales"
        },
        "purchases": {
            "fields": ["invoice_number", "supplier_name", "total", "discount", "payment_method", "status", "created_at", "note"],
            "label_ar": "المشتريات",
            "label_fr": "Purchases"
        },
        "expenses": {
            "fields": ["title", "amount", "category", "payment_method", "date", "notes", "recurring"],
            "label_ar": "المصاريف",
            "label_fr": "Expenses"
        },
        "debts": {
            "fields": ["customer_name", "amount", "remaining", "type", "status", "due_date", "created_at", "notes"],
            "label_ar": "الديون",
            "label_fr": "Debts"
        }
    }

    def _resolve_db(user, tenant_id):
        """Super admin can target a specific tenant's database via tenant_id."""
        if tenant_id and user.get("role") == "super_admin":
            return get_tenant_db(tenant_id)
        return db

    @router.get("/collections")
    async def get_exportable_collections(tenant_id: Optional[str] = None, user: dict = Depends(get_current_user)):
        """Get list of collections available for import/export"""
        tdb = _resolve_db(user, tenant_id)
        result = []
        for key, info in EXPORTABLE_COLLECTIONS.items():
            count = await tdb[key].count_documents({})
            result.append({
                "key": key,
                "label_ar": info["label_ar"],
                "label_fr": info["label_fr"],
                "fields": info["fields"],
                "count": count
            })
        return result

    @router.get("/export/{collection}")
    async def export_data(
        collection: str,
        format: str = Query("csv", enum=["csv", "xlsx"]),
        tenant_id: Optional[str] = None,
        user: dict = Depends(get_current_user)
    ):
        """Export collection data as CSV or Excel"""
        if collection not in EXPORTABLE_COLLECTIONS:
            raise HTTPException(status_code=400, detail=f"Collection '{collection}' not exportable")

        tdb = _resolve_db(user, tenant_id)
        fields = EXPORTABLE_COLLECTIONS[collection]["fields"]
        cursor = tdb[collection].find({}).sort("created_at", -1)
        docs = await cursor.to_list(length=50000)

        if format == "csv":
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=["id"] + fields, extrasaction='ignore')
            writer.writeheader()
            for doc in docs:
                row = {"id": str(doc.get("_id", ""))}
                for f in fields:
                    val = doc.get(f, "")
                    if isinstance(val, ObjectId):
                        val = str(val)
                    elif isinstance(val, datetime):
                        val = val.isoformat()
                    elif isinstance(val, (dict, list)):
                        val = json.dumps(val, ensure_ascii=False)
                    row[f] = val or ""
                writer.writerow(row)

            output.seek(0)
            filename = f"{collection}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )

        else:  # xlsx
            wb = Workbook()
            ws = wb.active
            ws.title = collection
            headers = ["id"] + fields
            ws.append(headers)

            for doc in docs:
                row = [str(doc.get("_id", ""))]
                for f in fields:
                    val = doc.get(f, "")
                    if isinstance(val, ObjectId):
                        val = str(val)
                    elif isinstance(val, datetime):
                        val = val.isoformat()
                    elif isinstance(val, (dict, list)):
                        val = json.dumps(val, ensure_ascii=False)
                    row.append(val or "")
                ws.append(row)

            # Auto-width columns
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"{collection}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )

    @router.post("/import/{collection}")
    async def import_data(
        collection: str,
        file: UploadFile = File(...),
        mode: str = Query("append", enum=["append", "replace"]),
        tenant_id: Optional[str] = Query(None),
        user: dict = Depends(get_current_user)
    ):
        """Import data from CSV or Excel file"""
        if collection not in EXPORTABLE_COLLECTIONS:
            raise HTTPException(status_code=400, detail=f"Collection '{collection}' not importable")

        if user.get("role") not in ["admin", "super_admin"]:
            raise HTTPException(status_code=403, detail="Admin only")

        tdb = _resolve_db(user, tenant_id)
        fields = EXPORTABLE_COLLECTIONS[collection]["fields"]
        content = await file.read()
        filename = file.filename.lower()
        records = []

        try:
            if filename.endswith('.csv'):
                text = content.decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(text))
                for row in reader:
                    record = {}
                    for f in fields:
                        if f in row and row[f]:
                            val = row[f]
                            # Type conversion
                            if f in ["retail_price", "wholesale_price", "purchase_price", "amount", "total", "discount", "salary", "remaining", "tax_rate", "quantity", "min_stock"]:
                                try:
                                    val = float(val)
                                except (ValueError, TypeError):
                                    val = 0
                            record[f] = val
                    record["created_at"] = datetime.now(timezone.utc).isoformat()
                    record["updated_at"] = datetime.now(timezone.utc).isoformat()
                    records.append(record)

            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = load_workbook(io.BytesIO(content), read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    raise HTTPException(status_code=400, detail="Empty file")

                headers_row = [str(h).strip() if h else "" for h in rows[0]]
                for row in rows[1:]:
                    record = {}
                    for i, val in enumerate(row):
                        if i < len(headers_row) and headers_row[i] in fields:
                            field = headers_row[i]
                            if val is None:
                                val = ""
                            if field in ["retail_price", "wholesale_price", "purchase_price", "amount", "total", "discount", "salary", "remaining", "tax_rate", "quantity", "min_stock"]:
                                try:
                                    val = float(val) if val else 0
                                except (ValueError, TypeError):
                                    val = 0
                            record[field] = val
                    record["created_at"] = datetime.now(timezone.utc).isoformat()
                    record["updated_at"] = datetime.now(timezone.utc).isoformat()
                    records.append(record)
                wb.close()
            else:
                raise HTTPException(status_code=400, detail="Unsupported file format. Use .csv or .xlsx")

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading file: {str(e)}")

        if not records:
            raise HTTPException(status_code=400, detail="No valid records found")

        # Log the import
        import_log = {
            "collection": collection,
            "filename": file.filename,
            "mode": mode,
            "records_count": len(records),
            "user_id": str(user.get("_id", "")),
            "user_name": user.get("name", ""),
            "created_at": datetime.now(timezone.utc).isoformat()
        }

        if mode == "replace":
            await tdb[collection].delete_many({})

        if records:
            result = await tdb[collection].insert_many(records)
            import_log["inserted_count"] = len(result.inserted_ids)

        await tdb["import_logs"].insert_one(import_log)

        return {
            "success": True,
            "message": f"Imported {len(records)} records to {collection}",
            "records_imported": len(records),
            "mode": mode
        }

    @router.get("/import-history")
    async def get_import_history(tenant_id: Optional[str] = None, user: dict = Depends(get_current_user)):
        """Get import history"""
        tdb = _resolve_db(user, tenant_id)
        cursor = tdb["import_logs"].find({}).sort("created_at", -1).limit(50)
        logs = await cursor.to_list(length=50)
        for log in logs:
            log["id"] = str(log.pop("_id"))
        return logs

    @router.get("/template/{collection}")
    async def download_template(
        collection: str,
        format: str = Query("csv", enum=["csv", "xlsx"]),
        user: dict = Depends(get_current_user)
    ):
        """Download empty template for import"""
        if collection not in EXPORTABLE_COLLECTIONS:
            raise HTTPException(status_code=400, detail=f"Invalid collection")

        fields = EXPORTABLE_COLLECTIONS[collection]["fields"]

        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(fields)
            # Add one sample row
            writer.writerow(["" for _ in fields])
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="{collection}_template.csv"'}
            )
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = collection
            ws.append(fields)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{collection}_template.xlsx"'}
            )

    return router
